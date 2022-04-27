import copy
from openpyxl import Workbook

class TestResultExcel(object):
    def __init__(self,sheet_name):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name
        self.row = 1
        self.datas = [] #所有输入的数据
        self.header = ["IMSI", "移动D", "移动E", "移动F","移动","电信B1","电信B3","电信","联通（B3）","总计", ]

        # 默认初始化表头
        for i in range(0,len(self.header)):
            self.ws.cell(row=1, column=i+1).value = self.header[i]
        self.row += 1

    # 重置表头信息
    def init_header(self, header):
        self.header = header
        # 更新表头显示
        for i in range(0, len(self.header)):
            self.ws.cell(row=1, column=i + 1).value = self.header[i]

    # 修改某一表头信息
    def alter_header(self, index, value):
        if index < len(self.header):
            self.header[index] = value
            self.ws.cell(row=1, column=index + 1).value = value

    def write_a_row(self,input_row_data):
        row_data = copy.deepcopy(input_row_data)
        sum = 0
        for i in range(1,len(row_data)):
            sum += row_data[i]
        row_data.append(sum)#捕获次数累计
        row_data.insert(4,row_data[1]+row_data[2]+row_data[3])#移动捕获总数
        row_data.insert(7,row_data[5]+row_data[6])#电信捕获总数

        for i in range(0,len(row_data)):
            self.ws.cell(row=self.row, column=i+1).value = row_data[i]
        self.row += 1
        self.datas.append(row_data)

    def save(self,file):
        footer_row = ["捕获率（%）"]
        for i in range(1,len(self.datas[0])):#表格列数
            exist_count = 0
            for d in range(0,len(self.datas)):#表格行数（不包括表头和表尾，只有中间数据部分）
                if self.datas[d][i] > 0:
                    exist_count += 1
            percentage = (exist_count*100) / len(self.datas)
            footer_row.append(percentage)
        #表最后一行写入捕获率
        for i in range(0,len(footer_row)):
            self.ws.cell(row=self.row, column=i+1).value = footer_row[i]
        self.wb.save(file)


def write_data_to_excel():
    nt = TestResultExcel("测试站点-时间")
    #header = ["IMSI","移动D","移动E","移动F","电信B1","电信B3","联通B3","累计",]
    #对应模块类型码        7       8       9       10      12      13
    #nt.set_header(header)

    imsi = ["9060",1,1,1,1,1,1]
    for i in range(0,10):
        nt.write_a_row(imsi)
        pass
    imsi = ["247612222233333", 2, 10, 0, 334, 1, 232]
    nt.write_a_row(imsi)
    imsi = ["247612222212333", 0, 0, 1, 0, 0, 0]
    nt.write_a_row(imsi)
    imsi = ["247612222212333", 0, 0, 0, 0, 0, 1]
    nt.write_a_row(imsi)

    nt.save("heihei.xlsx")